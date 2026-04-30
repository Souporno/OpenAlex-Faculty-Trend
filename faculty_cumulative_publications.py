"""
IO Faculty Master Publication Builder
======================================
Reads faculty roster from Excel (T1982–T2026 columns, 32 faculty),
matches each person to OpenAlex, fetches annual publication counts
for 1982–2026, and writes a master Excel with two sheets:

  Sheet 1 – Wide:  one row per faculty, columns Pub_YYYY & Cumul_YYYY
  Sheet 2 – Long:  one row per faculty-year (ready for Tableau CSV export)

Cumulative rule
---------------
  • Find first_title_year = first year the person holds a *faculty* title
    (Asst / Assoc / Full Prof).  PhD alone does NOT count.
  • Years BEFORE first_title_year  →  blank (None) for both Pub and Cumul
  • Years FROM first_title_year onward:
      Pub_YYYY   = actual publications that year
      Cumul_YYYY = running total of ALL papers since 1982
                   (pre-faculty papers are included in the running sum)

  Example: 6 papers published before becoming Asst Prof in 2005,
           3 papers in 2005  →  Cumul_2005 = 9.

Requirements
------------
    pip install requests openpyxl pandas tqdm

Usage
-----
    python faculty_cumulative_publications.py

    Update ROSTER_PATH to wherever your Excel file lives.
    Update EMAIL to your address (OpenAlex polite-pool requirement).
"""

import re
import time
import requests
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
EMAIL       = "ghoshsoupornoprawn@gmail.com"   # OpenAlex polite-pool
BASE_URL    = "https://api.openalex.org"
HEADERS     = {"User-Agent": f"mailto:{EMAIL}"}
YEARS       = list(range(1982, 2027))          # 1982–2026 inclusive
DELAY       = 0.15                             # seconds between API calls
ROSTER_PATH = "io_faculty_roster extended.xlsx"   # ← update path if needed
OUTPUT_PATH = "faculty_master_publications.xlsx"

# ── MANUALLY VERIFIED OPENALEX IDs ───────────────────────────────────────────
# Add entries here whenever auto-match is wrong.
# Format: "Faculty Name as in Excel": {"id": "A...", "matched_name": "...", "confidence": "HIGH"}
OPENALEX_OVERRIDES = {
    "Ze (Mia) Zhu": {
        "id": "A5084303359",
        "matched_name": "Ze Zhu",
        "confidence": "HIGH",
    },
}

# ── RANK PARSER ───────────────────────────────────────────────────────────────
def parse_rank(text):
    """
    Return (rank_label, rank_level) from a T-column cell.
    PhD alone is NOT a faculty title and returns (None, None).
    Handles typos ('Assitant'), multi-role strings ('Professor + Associate Chair'),
    and 'Distinguished' prefix.
    """
    if not text:
        return None, None
    # Take everything before ' - ' (university suffix)
    rank_section = str(text).split(' - ')[0].strip()
    # Split compound roles on ' + ' or ' and '
    parts = re.split(r' \+ | and ', rank_section, flags=re.IGNORECASE)
    for part in parts:
        p = part.strip().lower()
        # Skip bare PhD / doctoral student entries
        if re.fullmatch(r'ph\.?d\.?( student| candidate)?|doctoral student|graduate student', p):
            continue
        if 'assist' in p or 'assit' in p:        # catches typo "assitant"
            return 'Asst Prof', 1
        if 'assoc' in p and 'prof' in p:          # "associate professor" not "associate chair"
            return 'Assoc Prof', 2
        if 'prof' in p:                            # any remaining professor = Full
            return 'Full Prof', 3
    return None, None


# ── OPENALEX API HELPERS ──────────────────────────────────────────────────────
def api_get(endpoint, params=None):
    """Rate-limited GET to OpenAlex. Returns parsed JSON or None on error."""
    url = f"{BASE_URL}/{endpoint}"
    params = {**(params or {}), "mailto": EMAIL}
    time.sleep(DELAY)
    try:
        r = requests.get(url, params=params, headers=HEADERS, timeout=30)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"    [ERROR] {url}: {e}")
        return None


def find_author(name):
    """
    Search OpenAlex for an author by display name.
    Returns (openalex_id, matched_display_name, confidence) or (None, None, None).
    """
    data = api_get("authors", {
        "search": name,
        "per_page": 5,
        "select": "id,display_name,works_count,last_known_institutions",
    })
    if data and data.get("results"):
        best = data["results"][0]
        aid  = best["id"].split("/")[-1]
        return aid, best.get("display_name", ""), "AUTO"
    return None, None, None


def get_pubs_by_year(author_id):
    """
    Fetch publication counts per year for an author using OpenAlex group_by.
    Returns dict {year: count} for all years in YEARS (0 if no publications).
    One API call instead of one per year — much faster.
    """
    data = api_get("works", {
        "filter": f"author.id:{author_id}",
        "group_by": "publication_year",
        "per_page": 200,
    })
    year_counts = {yr: 0 for yr in YEARS}
    if data and data.get("group_by"):
        for entry in data["group_by"]:
            try:
                yr = int(entry["key"])
                if yr in year_counts:
                    year_counts[yr] = entry["count"]
            except (ValueError, KeyError):
                pass
    return year_counts


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    # ── Load roster ───────────────────────────────────────────────────────────
    wb_in   = openpyxl.load_workbook(ROSTER_PATH)
    ws_in   = wb_in.active
    headers = [str(c.value) if c.value is not None else '' for c in ws_in[1]]
    rows    = list(ws_in.iter_rows(min_row=2, values_only=True))

    T_COLS = {yr: headers.index(f'T{yr}')
              for yr in YEARS if f'T{yr}' in headers}

    # Deduplicate: keep first occurrence of each faculty name
    seen, faculty_list = set(), []
    for row in rows:
        name = row[headers.index('Faculty Name')]
        if name and name not in seen:
            seen.add(name)
            faculty_list.append(row)

    print(f"Faculty to process: {len(faculty_list)}")
    print(f"Years: {YEARS[0]}–{YEARS[-1]}  ({len(YEARS)} years)\n")

    wide_rows = []   # one dict per faculty  (Sheet 1)
    long_rows = []   # one dict per faculty-year (Sheet 2)

    for idx, row in enumerate(faculty_list):
        name        = row[headers.index('Faculty Name')]
        institution = row[headers.index('Institution')]
        abbrev      = row[headers.index('Abbreviation')]

        print(f"[{idx+1}/{len(faculty_list)}] {name}")

        # ── Determine first faculty title year ────────────────────────────────
        first_title_yr = None
        for yr in YEARS:
            if yr in T_COLS:
                rank_label, _ = parse_rank(row[T_COLS[yr]])
                if rank_label is not None:
                    first_title_yr = yr
                    break

        if first_title_yr is None:
            print(f"    WARNING: no faculty title found — skipping\n")
            continue
        print(f"    First title year: {first_title_yr}")

        # ── Resolve OpenAlex ID ───────────────────────────────────────────────
        if name in OPENALEX_OVERRIDES:
            ov            = OPENALEX_OVERRIDES[name]
            author_id     = ov["id"]
            matched_name  = ov["matched_name"]
            confidence    = ov["confidence"]
            print(f"    Override → {author_id} ({matched_name})")
        else:
            author_id, matched_name, confidence = find_author(name)
            if not author_id:
                print(f"    WARNING: not found on OpenAlex — skipping\n")
                continue
            print(f"    Matched  → {author_id} ({matched_name})")

        # ── Fetch annual publication counts (1 API call) ──────────────────────
        yearly_pubs = get_pubs_by_year(author_id)

        # ── Build cumulative (includes pre-faculty papers) ────────────────────
        running   = 0
        yearly_cumul = {}
        for yr in YEARS:
            running += yearly_pubs[yr]
            yearly_cumul[yr] = running

        total_all_time = running
        print(f"    Total pubs (all-time): {total_all_time}"
              f"  |  Cumul from {first_title_yr}: {yearly_cumul.get(2026, 0)}")

        # ── Wide row (one per faculty) ────────────────────────────────────────
        wide = {
            'Faculty Name':     name,
            'Institution':      institution,
            'Abbreviation':     abbrev,
            'OpenAlex ID':      author_id,
            'Matched Name':     matched_name,
            'Match Confidence': confidence,
            'First Title Year': first_title_yr,
            'Total All Time':   total_all_time,
        }
        for yr in YEARS:
            if yr < first_title_yr:
                wide[f'Pub_{yr}']   = None
                wide[f'Cumul_{yr}'] = None
            else:
                wide[f'Pub_{yr}']   = yearly_pubs[yr]
                wide[f'Cumul_{yr}'] = yearly_cumul[yr]
        wide_rows.append(wide)

        # ── Long rows (one per year per faculty) ──────────────────────────────
        for yr in YEARS:
            long_rows.append({
                'Faculty Name':       name,
                'Institution':        institution,
                'Abbreviation':       abbrev,
                'OpenAlex ID':        author_id,
                'Match Confidence':   confidence,
                'First Title Year':   first_title_yr,
                'Year':               yr,
                'Publications':       yearly_pubs[yr] if yr >= first_title_yr else None,
                'Cumul_Publications': yearly_cumul[yr] if yr >= first_title_yr else None,
            })

        print()

    # ── Export to Excel ───────────────────────────────────────────────────────
    print(f"Saving {OUTPUT_PATH} ...")

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    body_font = Font(name="Calibri", size=9)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb_out = Workbook()

    # ── Sheet 1: Wide ─────────────────────────────────────────────────────────
    ws1 = wb_out.active
    ws1.title = "Master (Wide)"

    fixed_cols = ['Faculty Name','Institution','Abbreviation',
                  'OpenAlex ID','Matched Name','Match Confidence',
                  'First Title Year','Total All Time']
    year_cols  = []
    for yr in YEARS:
        year_cols += [f'Pub_{yr}', f'Cumul_{yr}']
    all_cols = fixed_cols + year_cols

    for c, col in enumerate(all_cols, 1):
        cell = ws1.cell(row=1, column=c, value=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws1.column_dimensions[get_column_letter(c)].width = (
            20 if col in fixed_cols else 9)

    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for r, row_dict in enumerate(wide_rows, 2):
        fill = alt_fill if r % 2 == 0 else None
        for c, col in enumerate(all_cols, 1):
            cell = ws1.cell(row=r, column=c, value=row_dict.get(col))
            cell.font = body_font
            if fill:
                cell.fill = fill

    ws1.freeze_panes = "I2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(fixed_cols))}1"

    # ── Sheet 2: Long (Tableau-ready) ─────────────────────────────────────────
    ws2 = wb_out.create_sheet("Long (Tableau)")
    long_col_defs = [
        ('Faculty Name', 30), ('Institution', 35), ('Abbreviation', 12),
        ('OpenAlex ID', 16), ('Match Confidence', 18), ('First Title Year', 17),
        ('Year', 8), ('Publications', 14), ('Cumul_Publications', 20),
    ]
    for c, (col, width) in enumerate(long_col_defs, 1):
        cell = ws2.cell(row=1, column=c, value=col)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
        ws2.column_dimensions[get_column_letter(c)].width = width

    prev_name = None; alt = False
    for r, row_dict in enumerate(long_rows, 2):
        if row_dict['Faculty Name'] != prev_name:
            alt = not alt
            prev_name = row_dict['Faculty Name']
        fill = alt_fill if alt else None
        for c, (col, _) in enumerate(long_col_defs, 1):
            cell = ws2.cell(row=r, column=c, value=row_dict.get(col))
            cell.font = body_font
            if fill:
                cell.fill = fill

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    wb_out.save(OUTPUT_PATH)
    print(f"\n✓ Done.")
    print(f"  Faculty processed : {len(wide_rows)}")
    print(f"  Wide sheet rows   : {len(wide_rows)}")
    print(f"  Long sheet rows   : {len(long_rows)}")
    print(f"  Output            : {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
