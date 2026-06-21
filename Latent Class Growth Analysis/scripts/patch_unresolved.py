"""
patch_unresolved.py

Re-resolves the 4 unresolved authors using smarter URL parsing,
then re-fetches their publication counts and patches openalex_pub_counts.xlsx.

Run from the FullData folder:
    python3 patch_unresolved.py
"""

import re
import time
import requests
import openpyxl
from pathlib import Path

EMAIL    = "souporno@uw.edu"
BASE     = "https://api.openalex.org"
HEADERS  = {"User-Agent": f"mailto:{EMAIL}"}
YEARS    = list(range(1982, 2026))
ALYSSA_GAP = set(range(2015, 2019))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get(url, params):
    for attempt in range(4):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=20)
            if r.status_code == 429:
                time.sleep(2 ** attempt * 5)
                continue
            if r.status_code == 200:
                return r.json()
            print(f"  [HTTP {r.status_code}] {url} params={params}")
            return None
        except requests.RequestException as e:
            print(f"  [error] {e}")
            time.sleep(2 ** attempt)
    return None


def extract_doi_from_url(raw: str):
    """
    Extracts a usable DOI or identifier from various publisher URL formats.
    Returns (doi_or_id, id_type) where id_type is 'doi' or 'pmid' or 'pii'.
    """
    if not raw:
        return None, None
    parts = re.split(r"[;\n]", str(raw))
    for part in parts:
        part = part.strip()
        if not part:
            continue

        # Already a bare DOI or doi.org URL
        m = re.search(r"(?:https?://(?:dx\.)?doi\.org/|doi:)(.+)", part, re.IGNORECASE)
        if m:
            return m.group(1).strip(), "doi"

        # Tandfonline: .../doi/abs/{doi} or .../doi/full/{doi}
        m = re.search(r"tandfonline\.com/doi/(?:abs|full|pdf)/(.+)", part)
        if m:
            return m.group(1).strip(), "doi"

        # Annual Reviews: .../journals/{doi}
        m = re.search(r"annualreviews\.org/(?:content/journals|doi/abs)/(.+)", part)
        if m:
            return m.group(1).strip(), "doi"

        # Springer: link.springer.com/article/{doi}
        m = re.search(r"link\.springer\.com/article/(.+)", part)
        if m:
            return m.group(1).strip(), "doi"

        # PubMed: pubmed.ncbi.nlm.nih.gov/{pmid}
        m = re.search(r"pubmed\.ncbi\.nlm\.nih\.gov/(\d+)", part)
        if m:
            return m.group(1).strip(), "pmid"

        # ScienceDirect PII: .../pii/{pii}
        m = re.search(r"sciencedirect\.com/science/article/pii/([A-Z0-9]+)", part)
        if m:
            return m.group(1).strip(), "pii"

        # APA PsycNet: psycnet.apa.org/record/{doi}
        m = re.search(r"psycnet\.apa\.org/record/(.+)", part)
        if m:
            return m.group(1).strip(), "doi"

        # SAGE: journals.sagepub.com/doi/{type}/{doi}
        m = re.search(r"sagepub\.com/doi/(?:abs|full|pdf|10\.\d+/.+)", part)
        if m:
            doi = re.sub(r".*/doi/(?:abs|full|pdf)/", "", part).strip()
            if doi.startswith("10."):
                return doi, "doi"

        # Wiley: onlinelibrary.wiley.com/doi/{doi}
        m = re.search(r"onlinelibrary\.wiley\.com/doi/(.+)", part)
        if m:
            return m.group(1).strip(), "doi"

        # Generic: if it starts with 10. it's probably a bare DOI
        if part.startswith("10."):
            return part, "doi"

    return None, None


def find_author_via_doi(doi: str, faculty_name: str):
    data = get(f"{BASE}/works", {"filter": f"doi:{doi}", "mailto": EMAIL})
    if not data or not data.get("results"):
        return None, "NOT_FOUND"
    work = data["results"][0]
    return _match_author(work.get("authorships", []), faculty_name)


def find_author_via_pmid(pmid: str, faculty_name: str):
    data = get(f"{BASE}/works", {"filter": f"ids.pmid:{pmid}", "mailto": EMAIL})
    if not data or not data.get("results"):
        return None, "NOT_FOUND"
    work = data["results"][0]
    return _match_author(work.get("authorships", []), faculty_name)


def find_author_via_pii(pii: str, faculty_name: str):
    """Try OpenAlex landing-page filter for ScienceDirect PII."""
    landing = f"https://www.sciencedirect.com/science/article/pii/{pii}"
    data = get(f"{BASE}/works",
               {"filter": f"locations.landing_page_url:{landing}", "mailto": EMAIL})
    if data and data.get("results"):
        work = data["results"][0]
        result = _match_author(work.get("authorships", []), faculty_name)
        if result[0]:
            return result
    # Fallback: search by author name
    return find_author_by_name(faculty_name)


def find_author_by_name(name: str):
    data = get(f"{BASE}/authors", {"search": name, "mailto": EMAIL})
    if not data or not data.get("results"):
        return None, "NOT_FOUND"
    norm = lambda s: re.sub(r"\s+", " ", s.lower().strip())
    name_norm = norm(name)
    name_parts = set(name_norm.split())
    for r in data["results"][:5]:
        cand_norm = norm(r.get("display_name", ""))
        if cand_norm == name_norm:
            return r["id"], "HIGH"
    for r in data["results"][:5]:
        cand_parts = set(norm(r.get("display_name", "")).split())
        score = len(name_parts & cand_parts) / max(len(name_parts), len(cand_parts))
        if score >= 0.5:
            return r["id"], "MEDIUM"
    return None, "NOT_FOUND"


def _match_author(authorships, faculty_name: str):
    norm = lambda s: re.sub(r"\s+", " ", s.lower().strip())
    fn = norm(faculty_name)
    fp = set(fn.split())
    best_id, best_conf, best_score = None, "NOT_FOUND", 0
    for auth in authorships:
        author = auth.get("author", {})
        aid    = author.get("id", "")
        aname  = norm(author.get("display_name", ""))
        if aname == fn:
            return aid, "HIGH"
        ap = set(aname.split())
        score = len(fp & ap) / max(len(fp), len(ap))
        if score > best_score:
            best_score, best_id = score, aid
            best_conf = "HIGH" if score >= 0.8 else "MEDIUM" if score >= 0.5 else "LOW"
    return best_id, best_conf


def get_pub_count(author_id: str, year: int) -> int:
    short_id = author_id.split("/")[-1]
    data = get(f"{BASE}/works",
               {"filter": f"author.id:{short_id},publication_year:{year}",
                "per_page": 1, "mailto": EMAIL})
    return data.get("meta", {}).get("count", 0) if data else 0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    input_xl  = Path("Encoded.xlsx")
    output_xl = Path("openalex_pub_counts.xlsx")

    if not input_xl.exists():
        print("ERROR: 'Encoded io_faculty.xlsx' not found.")
        return
    if not output_xl.exists():
        print("ERROR: 'openalex_pub_counts.xlsx' not found. Run fetch_openalex_pubs.py first.")
        return

    # Read source Excel for positions + DOIs of unresolved
    src_wb = openpyxl.load_workbook(input_xl)
    src_ws = src_wb.active
    headers = [cell.value for cell in src_ws[1]]

    col_name = headers.index("Faculty Name")
    col_doi  = headers.index("Publication(s) for ID")
    year_col = {y: headers.index(y) for y in YEARS if y in headers}

    unresolved_faculty = {}
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        name = str(row[col_name]).strip() if row[col_name] else ""
        if not name:
            continue
        positions = {}
        for y, ci in year_col.items():
            val = row[ci]
            if val is not None and str(val).strip() != "":
                try:
                    positions[y] = int(val)
                except (ValueError, TypeError):
                    pass
        unresolved_faculty[name] = {
            "doi_raw":   row[col_doi],
            "positions": positions,
        }

    # Read output Excel — find NOT_FOUND rows
    out_wb = openpyxl.load_workbook(output_xl)
    out_ws = out_wb.active
    out_headers = [cell.value for cell in out_ws[1]]
    col_aid  = out_headers.index("OpenAlex Author ID")
    col_conf = out_headers.index("Match Confidence")

    to_patch = []   # (row_idx_1based, name)
    for i, row in enumerate(out_ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[col_conf] == "NOT_FOUND":
            to_patch.append((i, str(row[0]).strip()))

    if not to_patch:
        print("No NOT_FOUND rows found in output — nothing to patch.")
        return

    print(f"Found {len(to_patch)} NOT_FOUND rows: {[n for _, n in to_patch]}\n")

    # Resolve each
    for row_idx, name in to_patch:
        print(f"Resolving: {name}")
        fac = unresolved_faculty.get(name, {})
        doi_raw   = fac.get("doi_raw", "")
        positions = fac.get("positions", {})

        identifier, id_type = extract_doi_from_url(doi_raw)
        author_id, confidence = None, "NOT_FOUND"

        if identifier:
            print(f"  Identifier: {identifier!r} (type={id_type})")
            if id_type == "doi":
                author_id, confidence = find_author_via_doi(identifier, name)
            elif id_type == "pmid":
                author_id, confidence = find_author_via_pmid(identifier, name)
            elif id_type == "pii":
                author_id, confidence = find_author_via_pii(identifier, name)

        if not author_id:
            print(f"  DOI/ID lookup failed, trying name search …")
            author_id, confidence = find_author_by_name(name)

        print(f"  → {author_id} [{confidence}]")

        if not author_id:
            print(f"  Still unresolved — skipping.\n")
            continue

        # Fetch publication counts
        active_years = sorted(positions.keys())
        is_alyssa = name.lower() == "alyssa gibbons"
        pub_counts, cumulative = {}, {}

        if is_alyssa:
            for run in [[y for y in active_years if y < min(ALYSSA_GAP)],
                        [y for y in active_years if y > max(ALYSSA_GAP)]]:
                cum = 0
                for year in run:
                    cnt = get_pub_count(author_id, year)
                    pub_counts[year] = cnt
                    cum += cnt
                    cumulative[year] = cum
                    time.sleep(0.12)
        else:
            cum = 0
            for year in active_years:
                cnt = get_pub_count(author_id, year)
                pub_counts[year] = cnt
                cum += cnt
                cumulative[year] = cum
                time.sleep(0.12)

        # Patch output row
        row_obj = out_ws[row_idx]
        row_obj[col_aid].value  = author_id
        row_obj[col_conf].value = confidence

        for y in YEARS:
            col_count = out_headers.index(f"{y}_count")
            col_cum   = out_headers.index(f"{y}_cumulative")
            if y in positions:
                row_obj[col_count].value = pub_counts.get(y, "")
                row_obj[col_cum].value   = cumulative.get(y, "")

        print(f"  Patched row {row_idx}.\n")
        time.sleep(0.2)

    out_wb.save(output_xl)
    print(f"Done! Patched file saved to '{output_xl}'.")


if __name__ == "__main__":
    main()
