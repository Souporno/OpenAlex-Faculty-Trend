"""
fetch_openalex_pubs.py

Reads 'Encoded io_faculty.xlsx', resolves each faculty member's OpenAlex
Author ID (using the DOI in 'Publication(s) for ID'), then fetches annual
publication counts for every year the faculty held a position (encoded in
the 1982-2025 integer columns).

Outputs
-------
1. openalex_pub_counts.xlsx  – one sheet with:
     Faculty Name | Author ID | Match Confidence |
     {year}_count  (pub count that year)
     {year}_cumulative  (running cumulative)
   Only years where the faculty held a position are populated.

2. unresolved_authors.txt  – names for which no Author ID was found,
   so you can supply a different DOI.

Usage
-----
    python fetch_openalex_pubs.py

Requirements
------------
    pip install openpyxl requests tqdm
"""

import re
import sys
import time
import json
import requests
import openpyxl
from tqdm import tqdm
from pathlib import Path

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_FILE   = "Encoded.xlsx"
OUTPUT_FILE  = "openalex_pub_counts.xlsx"
UNRESOLVED   = "unresolved_authors.txt"
EMAIL        = "souporno@uw.edu"           # polite pool
BASE         = "https://api.openalex.org"
YEARS        = list(range(1982, 2026))     # 1982-2025 inclusive
ALYSSA_NAME  = "Alyssa Gibbons"
# Alyssa gap: she held no position 2015-2018; treat as two continuous runs.
ALYSSA_GAP   = set(range(2015, 2019))      # 2015,2016,2017,2018

HEADERS_GET  = {"User-Agent": f"mailto:{EMAIL}"}


def clean_doi(raw: str):
    """Extract a single clean DOI from a (possibly multi-value) cell string."""
    if not raw:
        return None
    # split on semicolons or newlines, take the first non-empty chunk
    parts = re.split(r"[;\n]", str(raw))
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # normalise to bare DOI
        part = re.sub(r"https?://(dx\.)?doi\.org/", "", part)
        part = re.sub(r"^doi:", "", part, flags=re.IGNORECASE).strip()
        if part:
            return part
    return None


def get_with_retry(url: str, params: dict, retries: int = 4):
    """GET with exponential back-off; returns JSON dict or None on failure."""
    for attempt in range(retries):
        try:
            r = requests.get(url, params=params, headers=HEADERS_GET, timeout=20)
            if r.status_code == 429:
                wait = 2 ** attempt * 5
                print(f"  [429] rate-limited, sleeping {wait}s …")
                time.sleep(wait)
                continue
            if r.status_code == 200:
                return r.json()
            print(f"  [HTTP {r.status_code}] {url}")
            return None
        except requests.RequestException as e:
            print(f"  [error] {e}")
            time.sleep(2 ** attempt)
    return None


# ---------------------------------------------------------------------------
# Author resolution
# ---------------------------------------------------------------------------

def resolve_author_via_doi(doi: str, faculty_name: str):
    """
    Look up a work by DOI, find the authorship that best matches faculty_name.
    Returns (openalex_author_id, confidence).
    confidence ∈ {"HIGH", "MEDIUM", "LOW", "NOT_FOUND"}
    """
    doi_clean = doi.strip()
    data = get_with_retry(f"{BASE}/works", {"filter": f"doi:{doi_clean}", "mailto": EMAIL})
    if not data or data.get("meta", {}).get("count", 0) == 0:
        return None, "NOT_FOUND"

    works = data.get("results", [])
    if not works:
        return None, "NOT_FOUND"

    work = works[0]
    authorships = work.get("authorships", [])
    if not authorships:
        return None, "NOT_FOUND"

    # Normalise name for comparison
    def norm(s):
        return re.sub(r"\s+", " ", s.lower().strip())

    faculty_norm = norm(faculty_name)
    faculty_parts = set(faculty_norm.split())

    best_id   = None
    best_conf = "NOT_FOUND"
    best_score = 0

    for auth in authorships:
        author = auth.get("author", {})
        aid    = author.get("id", "")
        aname  = author.get("display_name", "")
        aname_norm = norm(aname)

        # Exact match
        if aname_norm == faculty_norm:
            return aid, "HIGH"

        # Token overlap
        aparts = set(aname_norm.split())
        overlap = len(faculty_parts & aparts)
        score   = overlap / max(len(faculty_parts), len(aparts))

        if score > best_score:
            best_score = score
            best_id = aid
            if score >= 0.8:
                best_conf = "HIGH"
            elif score >= 0.5:
                best_conf = "MEDIUM"
            else:
                best_conf = "LOW"

    return best_id, best_conf


def resolve_author_by_name(name: str):
    """
    Search OpenAlex /authors by display_name for faculty with no DOI.
    Returns (openalex_author_id, confidence).
    """
    data = get_with_retry(f"{BASE}/authors",
                          {"search": name, "mailto": EMAIL})
    if not data or not data.get("results"):
        return None, "NOT_FOUND"

    results = data["results"]
    norm = lambda s: re.sub(r"\s+", " ", s.lower().strip())
    name_norm = norm(name)

    for r in results[:5]:
        cand_norm = norm(r.get("display_name", ""))
        if cand_norm == name_norm:
            return r["id"], "HIGH"

    # Fallback: first result with token overlap >= 0.5
    name_parts = set(name_norm.split())
    for r in results[:5]:
        cand_parts = set(norm(r.get("display_name", "")).split())
        score = len(name_parts & cand_parts) / max(len(name_parts), len(cand_parts))
        if score >= 0.5:
            return r["id"], "MEDIUM"

    return None, "NOT_FOUND"


# ---------------------------------------------------------------------------
# Publication counts
# ---------------------------------------------------------------------------

def get_pub_count_for_year(author_id: str, year: int) -> int:
    """
    Returns the number of works authored by author_id in the given year.
    Uses per_page=1 because we only need the meta count.
    """
    short_id = author_id.split("/")[-1]   # e.g. A123456789
    data = get_with_retry(
        f"{BASE}/works",
        {
            "filter": f"author.id:{short_id},publication_year:{year}",
            "per_page": 1,
            "mailto": EMAIL,
        },
    )
    if data is None:
        return 0
    return data.get("meta", {}).get("count", 0)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        sys.exit(f"ERROR: '{EXCEL_FILE}' not found in current directory.")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Column indices (0-based)
    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    col_name = col("Faculty Name")
    col_doi  = col("Publication(s) for ID")

    # Map year -> column index for position encoding
    year_col = {}
    for y in YEARS:
        idx = col(y)          # headers contain int years
        if idx is not None:
            year_col[y] = idx

    # Read all faculty rows (skip header)
    faculty_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col_name] if col_name is not None else None
        if not name or str(name).strip() == "":
            continue
        name = str(name).strip()
        doi_raw = row[col_doi] if col_doi is not None else None
        positions = {}
        for y, ci in year_col.items():
            val = row[ci]
            if val is not None and str(val).strip() != "":
                try:
                    positions[y] = int(val)
                except (ValueError, TypeError):
                    pass
        faculty_rows.append({"name": name, "doi_raw": doi_raw, "positions": positions})

    print(f"Loaded {len(faculty_rows)} faculty rows.\n")

    # ------------------------------------------------------------------
    # Phase 1: Resolve Author IDs
    # ------------------------------------------------------------------
    print("=" * 60)
    print("Phase 1: Resolving OpenAlex Author IDs …")
    print("=" * 60)

    resolved   = []   # list of dicts with name, author_id, confidence, positions
    unresolved = []   # names with no author_id found

    for fac in tqdm(faculty_rows, unit="faculty"):
        name     = fac["name"]
        doi_raw  = fac["doi_raw"]
        doi      = clean_doi(doi_raw)
        author_id = None
        confidence = "NOT_FOUND"

        if doi:
            author_id, confidence = resolve_author_via_doi(doi, name)

        # Fallback to name search (also used for John Schmidt)
        if not author_id:
            author_id, confidence = resolve_author_by_name(name)

        if not author_id:
            unresolved.append(name)
            print(f"  [UNRESOLVED] {name}")

        resolved.append({
            "name":       name,
            "author_id":  author_id or "",
            "confidence": confidence,
            "positions":  fac["positions"],
        })

        time.sleep(0.12)   # ~8 req/s — well within polite pool

    print(f"\nResolved: {len(resolved) - len(unresolved)} / {len(resolved)}")
    print(f"Unresolved: {len(unresolved)}\n")

    # Write unresolved list
    with open(UNRESOLVED, "w") as f:
        for name in unresolved:
            f.write(name + "\n")
    if unresolved:
        print(f"Unresolved names written to '{UNRESOLVED}'.\n")

    # ------------------------------------------------------------------
    # Phase 2: Fetch publication counts
    # ------------------------------------------------------------------
    print("=" * 60)
    print("Phase 2: Fetching annual publication counts …")
    print("=" * 60)

    for fac in tqdm(resolved, unit="faculty"):
        if not fac["author_id"]:
            fac["pub_counts"] = {}
            continue

        positions = fac["positions"]
        active_years = sorted(positions.keys())

        # Special handling for Alyssa Gibbons (discontinuous)
        is_alyssa = fac["name"].lower() == ALYSSA_NAME.lower()

        pub_counts   = {}   # year -> count that year
        cumulative   = {}   # year -> cumulative up to and including year

        if is_alyssa:
            # Split into two continuous runs: before gap and after gap
            run_before = [y for y in active_years if y < min(ALYSSA_GAP)]
            run_after  = [y for y in active_years if y > max(ALYSSA_GAP)]

            for run in [run_before, run_after]:
                cum = 0
                for year in run:
                    cnt = get_pub_count_for_year(fac["author_id"], year)
                    pub_counts[year] = cnt
                    cum += cnt
                    cumulative[year] = cum
                    time.sleep(0.12)
        else:
            cum = 0
            for year in active_years:
                cnt = get_pub_count_for_year(fac["author_id"], year)
                pub_counts[year] = cnt
                cum += cnt
                cumulative[year] = cum
                time.sleep(0.12)

        fac["pub_counts"] = pub_counts
        fac["cumulative"] = cumulative

    # ------------------------------------------------------------------
    # Phase 3: Write output Excel
    # ------------------------------------------------------------------
    print("\nWriting output …")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Publication Counts"

    # Build header row
    header = ["Faculty Name", "OpenAlex Author ID", "Match Confidence"]
    for y in YEARS:
        header.append(f"{y}_count")
        header.append(f"{y}_cumulative")
    out_ws.append(header)

    for fac in resolved:
        row = [fac["name"], fac["author_id"], fac["confidence"]]
        pub_counts = fac.get("pub_counts", {})
        cumulative = fac.get("cumulative", {})
        positions  = fac["positions"]

        for y in YEARS:
            if y in positions:
                row.append(pub_counts.get(y, ""))
                row.append(cumulative.get(y, ""))
            else:
                row.append("")
                row.append("")
        out_ws.append(row)

    # Light formatting: freeze top row
    out_ws.freeze_panes = "A2"
    out_wb.save(OUTPUT_FILE)
    print(f"Done! Results saved to '{OUTPUT_FILE}'.")


if __name__ == "__main__":
    main()
