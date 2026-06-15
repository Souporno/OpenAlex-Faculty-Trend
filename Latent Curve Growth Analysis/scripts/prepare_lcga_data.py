"""
prepare_lcga_data.py

Merges Encoded.xlsx and openalex_pub_counts.xlsx into a long-format CSV
suitable for LCGA in R (lcmm package).

Output: lcga_data_long.csv

Columns in output:
  faculty_id          : integer ID (1–N)
  faculty_name        : string
  university          : current institution
  calendar_year       : actual year (e.g. 2005)
  career_year         : years since career start, 1-indexed
  pub_count           : publications that calendar year (from OpenAlex)
  cumulative_pub      : cumulative publications up to and including that year
  rank                : 1=Asst, 2=Assoc, 3=Full
  promoted_this_year  : 1 if rank increased from previous year, else 0
  career_start_year   : calendar year of career year 1
  never_asst_prof     : 1 if faculty never held rank=1 (Asst Prof)
  author_id           : OpenAlex author ID
  match_confidence    : HIGH / MEDIUM / LOW / NOT_FOUND

Usage:
    python3 prepare_lcga_data.py
"""

import openpyxl
import csv
from pathlib import Path

ENCODED_FILE    = "Encoded.xlsx"
PUB_COUNT_FILE  = "openalex_pub_counts.xlsx"
OUTPUT_FILE     = "lcga_data_long.csv"
YEARS           = list(range(1982, 2026))   # 1982-2025
MIN_OBS         = 3    # minimum years with a position to include in LCGA


def load_encoded():
    """Return list of dicts, one per faculty, from Encoded.xlsx."""
    wb = openpyxl.load_workbook(ENCODED_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    col = lambda name: headers.index(name)

    col_name  = col("Faculty Name")
    col_uni   = col("University Name")

    # Integer year columns
    year_cols = {}
    for y in YEARS:
        if y in headers:
            year_cols[y] = headers.index(y)

    faculty = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[col_name]).strip() if row[col_name] else ""
        if not name:
            continue
        uni = str(row[col_uni]).strip() if row[col_uni] else ""

        positions = {}
        for y, ci in year_cols.items():
            val = row[ci]
            if val is not None and str(val).strip() != "":
                try:
                    positions[y] = int(val)
                except (ValueError, TypeError):
                    pass

        faculty.append({
            "name":      name,
            "university": uni,
            "positions": positions,   # {calendar_year: rank}
        })

    return faculty


def load_pub_counts():
    """Return dict: faculty_name -> {year -> {count, cumulative}}."""
    wb = openpyxl.load_workbook(PUB_COUNT_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    pub = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if not name:
            continue
        author_id  = row[1] if len(row) > 1 else ""
        confidence = row[2] if len(row) > 2 else ""

        year_data = {}
        for y in YEARS:
            try:
                col_count = headers.index(f"{y}_count")
                col_cum   = headers.index(f"{y}_cumulative")
                cnt = row[col_count]
                cum = row[col_cum]
                year_data[y] = {
                    "count":      int(cnt) if cnt not in (None, "") else None,
                    "cumulative": int(cum) if cum not in (None, "") else None,
                }
            except (ValueError, IndexError):
                year_data[y] = {"count": None, "cumulative": None}

        pub[name] = {
            "author_id":  author_id,
            "confidence": confidence,
            "years":      year_data,
        }

    return pub


def derive_career_year(positions):
    """
    Returns (career_start_year, never_asst_prof) given a {year: rank} dict.

    career_start_year = first year with rank == 1 (Asst Prof).
    If no rank-1 years exist, fall back to first year with any rank.
    never_asst_prof   = True if no rank-1 years in the data.
    """
    asst_years = sorted(y for y, r in positions.items() if r == 1)
    if asst_years:
        return asst_years[0], False
    # Fallback: first observed year regardless of rank
    all_years = sorted(positions.keys())
    return (all_years[0], True) if all_years else (None, True)


def main():
    print(f"Reading {ENCODED_FILE} …")
    faculty_list = load_encoded()
    print(f"  {len(faculty_list)} faculty loaded.")

    print(f"Reading {PUB_COUNT_FILE} …")
    pub_data = load_pub_counts()
    print(f"  {len(pub_data)} faculty with pub count data.")

    # Match names (exact)
    unmatched_in_pub = set(pub_data.keys()) - {f["name"] for f in faculty_list}
    if unmatched_in_pub:
        print(f"  Warning: {len(unmatched_in_pub)} names in pub file not in Encoded.xlsx")

    rows = []
    excluded_no_id  = []
    excluded_short  = []
    never_asst_count = 0

    for fid, fac in enumerate(faculty_list, start=1):
        name = fac["name"]
        uni  = fac["university"]
        positions = fac["positions"]

        if not positions:
            excluded_short.append(name)
            continue

        # Career year derivation
        career_start, never_asst = derive_career_year(positions)
        if career_start is None:
            excluded_short.append(name)
            continue
        if never_asst:
            never_asst_count += 1

        # Get pub data
        pub_entry = pub_data.get(name, {})
        author_id  = pub_entry.get("author_id", "")
        confidence = pub_entry.get("confidence", "NOT_FOUND")
        year_pub   = pub_entry.get("years", {})

        # Exclude if no author ID resolved
        if not author_id or confidence == "NOT_FOUND":
            excluded_no_id.append(name)
            continue

        # Build active years (years holding a position, sorted)
        active_years = sorted(positions.keys())

        if len(active_years) < MIN_OBS:
            excluded_short.append(name)
            continue

        # Build promotion flags: compare rank to previous active year
        sorted_active = sorted(active_years)
        rank_by_year  = {y: positions[y] for y in sorted_active}
        promoted_flag = {}
        for i, y in enumerate(sorted_active):
            if i == 0:
                promoted_flag[y] = 0
            else:
                prev_y = sorted_active[i - 1]
                promoted_flag[y] = 1 if rank_by_year[y] > rank_by_year[prev_y] else 0

        for y in sorted_active:
            career_year = y - career_start + 1
            pub_yr      = year_pub.get(y, {})
            count       = pub_yr.get("count", None)
            cumulative  = pub_yr.get("cumulative", None)

            rows.append({
                "faculty_id":         fid,
                "faculty_name":       name,
                "university":         uni,
                "calendar_year":      y,
                "career_year":        career_year,
                "pub_count":          count if count is not None else "",
                "cumulative_pub":     cumulative if cumulative is not None else "",
                "rank":               positions[y],
                "promoted_this_year": promoted_flag[y],
                "career_start_year":  career_start,
                "never_asst_prof":    int(never_asst),
                "author_id":          author_id,
                "match_confidence":   confidence,
            })

    # Write CSV
    fieldnames = [
        "faculty_id", "faculty_name", "university",
        "calendar_year", "career_year",
        "pub_count", "cumulative_pub",
        "rank", "promoted_this_year",
        "career_start_year", "never_asst_prof",
        "author_id", "match_confidence",
    ]
    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nOutput written to '{OUTPUT_FILE}'.")
    print(f"  Total row count : {len(rows)}")
    print(f"  Faculty included: {len(set(r['faculty_id'] for r in rows))}")
    print(f"  Excluded — no OpenAlex ID : {len(excluded_no_id)}")
    if excluded_no_id:
        for n in excluded_no_id:
            print(f"    - {n}")
    print(f"  Excluded — < {MIN_OBS} observed years : {len(excluded_short)}")
    if excluded_short:
        for n in excluded_short:
            print(f"    - {n}")
    print(f"  Never held Asst Prof rank (flagged, not excluded): {never_asst_count}")


if __name__ == "__main__":
    main()
