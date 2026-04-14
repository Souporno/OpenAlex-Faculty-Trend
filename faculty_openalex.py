#!/usr/bin/env python3
"""
OpenAlex Faculty Publication Counter
Reads 34 I/O faculty from Excel, queries OpenAlex for publication counts 2014-2026.
Outputs a new Excel file with per-year counts and match confidence flags.

Usage:
    pip install requests pandas openpyxl
    python faculty_openalex.py
"""

import re
import sys
import time

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
INPUT_FILE  = "Faculty Details Completed.xlsx"
OUTPUT_FILE = "faculty_publications_openalex.xlsx"
EMAIL       = "souporno@uw.edu"   # OpenAlex polite pool (faster rate limits)
YEARS       = list(range(2014, 2027))
BASE_URL    = "https://api.openalex.org"
PAUSE       = 0.35               # seconds between requests
# ─────────────────────────────────────────────────────────────────────────────


def clean_name(name: str) -> str:
    """Remove parenthetical nicknames, Jr./Sr., extra spaces."""
    name = re.sub(r'\s*\([^)]*\)', '', name)          # remove (Daisy), (Mia) etc.
    name = re.sub(r',?\s*(Jr\.|Sr\.|III|II)$', '', name, flags=re.IGNORECASE)
    return re.sub(r'\s+', ' ', name).strip()


def inst_keywords(institution: str) -> set:
    """Extract meaningful words from institution name for matching."""
    stop = {"university", "institute", "technology", "college", "state",
            "the", "of", "and", "at", "for", "georgia", "new"}
    words = set(w.lower() for w in institution.split() if len(w) > 2)
    return words - stop


def score_author(author: dict, target_name: str, inst_kws: set) -> int:
    """Score an OpenAlex author result against our target."""
    score = 0

    # ── Name match ───────────────────────────────────────────────────────────
    oa_name = author.get("display_name", "").lower()
    t_lower = target_name.lower()
    if oa_name == t_lower:
        score += 20
    elif t_lower in oa_name or oa_name in t_lower:
        score += 12
    else:
        last = t_lower.split()[-1]
        if last in oa_name:
            score += 5

    # ── Institution match ────────────────────────────────────────────────────
    all_insts = list(author.get("last_known_institutions") or [])
    for aff in author.get("affiliations") or []:
        obj = aff.get("institution") or {}
        if obj:
            all_insts.append(obj)

    for inst in all_insts:
        inst_name = (inst.get("display_name") or "").lower()
        hits = sum(1 for kw in inst_kws if kw in inst_name)
        if hits:
            score += 10 * hits
            break

    return score


def search_author(name: str, institution: str):
    """Return (best_author_dict, confidence_str) from OpenAlex."""
    query = clean_name(name)
    kws   = inst_keywords(institution)

    params = {
        "search":   query,
        "mailto":   EMAIL,
        "per_page": 10,
        "select":   "id,display_name,works_count,last_known_institutions,affiliations",
    }
    try:
        r = requests.get(f"{BASE_URL}/authors", params=params, timeout=12)
        r.raise_for_status()
        results = r.json().get("results", [])
    except Exception as e:
        print(f"    ⚠ API error: {e}")
        return None, "API_ERROR"

    if not results:
        return None, "NOT_FOUND"

    scored = [(score_author(a, query, kws), a) for a in results]
    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, best = scored[0]

    if best_score >= 20:
        confidence = "HIGH"
    elif best_score >= 10:
        confidence = "MEDIUM"
    elif best_score >= 5:
        confidence = "LOW"
    else:
        return None, "NOT_FOUND"

    return best, confidence


def get_year_counts(author_id: str) -> dict:
    """Return {year: count} from OpenAlex works grouped by publication_year."""
    params = {
        "filter":   f"author.id:{author_id}",
        "group_by": "publication_year",
        "mailto":   EMAIL,
        "per_page": 200,
    }
    try:
        r = requests.get(f"{BASE_URL}/works", params=params, timeout=12)
        r.raise_for_status()
        groups = r.json().get("group_by", [])
    except Exception as e:
        print(f"    ⚠ API error (works): {e}")
        return {}

    counts = {}
    for g in groups:
        try:
            counts[int(g["key"])] = g["count"]
        except (ValueError, KeyError):
            pass
    return counts


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 62)
    print("  OpenAlex Faculty Publication Counter  (2014 – 2026)")
    print("=" * 62)

    df = pd.read_excel(INPUT_FILE)
    print(f"  Loaded {len(df)} faculty from '{INPUT_FILE}'\n")

    rows = []

    for i, row in df.iterrows():
        name        = str(row["Faculty Name"]).strip()
        institution = str(row["Institution"]).strip()

        print(f"[{i+1:2}/{len(df)}] {name}  —  {institution}")

        author, confidence = search_author(name, institution)
        time.sleep(PAUSE)

        if author is None:
            print(f"       ✗  {confidence}")
            entry = {
                "Institution":          institution,
                "Faculty Name":         name,
                "OpenAlex ID":          confidence,   # e.g. "NOT_FOUND"
                "OpenAlex Matched Name": "",
                "Match Confidence":     confidence,
                "Total Works":          "",
            }
            for yr in YEARS:
                entry[str(yr)] = ""
        else:
            oa_id    = author["id"].split("/")[-1]
            oa_name  = author.get("display_name", "")
            total    = author.get("works_count", 0)

            yr_counts = get_year_counts(author["id"])
            time.sleep(PAUSE)

            icon = {"HIGH": "✓", "MEDIUM": "~", "LOW": "?"}.get(confidence, "?")
            print(f"       {icon}  {oa_name}  |  {oa_id}  |  {confidence}  |  {total} total works")

            entry = {
                "Institution":           institution,
                "Faculty Name":          name,
                "OpenAlex ID":           oa_id,
                "OpenAlex Matched Name": oa_name,
                "Match Confidence":      confidence,
                "Total Works":           total,
            }
            for yr in YEARS:
                entry[str(yr)] = yr_counts.get(yr, 0)

        rows.append(entry)

    # ── Build Excel ───────────────────────────────────────────────────────────
    print(f"\n  Writing '{OUTPUT_FILE}' ...")

    out_df  = pd.DataFrame(rows)
    headers = list(out_df.columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Publication Counts"

    DARK_BLUE  = "1F4E79"
    MED_BLUE   = "2E75B6"
    YELLOW     = "FFF2CC"
    LIGHT_BLUE = "DEEAF1"
    ORANGE     = "FCE4D6"
    WHITE      = "FFFFFF"

    # ── Header row ───────────────────────────────────────────────────────────
    for col_idx, h in enumerate(headers, 1):
        is_year = h in [str(y) for y in YEARS]
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill  = PatternFill("solid", start_color=MED_BLUE if is_year else DARK_BLUE)
        cell.font  = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, (_, data_row) in enumerate(out_df.iterrows(), 2):
        conf = data_row.get("Match Confidence", "")
        oa   = data_row.get("OpenAlex ID", "")

        if oa in ("NOT_FOUND", "API_ERROR"):
            row_fill = PatternFill("solid", start_color=ORANGE)
        elif conf == "LOW":
            row_fill = PatternFill("solid", start_color=YELLOW)
        elif conf == "MEDIUM":
            row_fill = PatternFill("solid", start_color=LIGHT_BLUE)
        else:
            row_fill = None

        for col_idx, h in enumerate(headers, 1):
            val  = data_row[h]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(
                horizontal="center" if h in [str(y) for y in YEARS] + ["Total Works", "Match Confidence"] else "left"
            )
            if row_fill:
                cell.fill = row_fill

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {
        "Institution": 30, "Faculty Name": 26, "OpenAlex ID": 16,
        "OpenAlex Matched Name": 30, "Match Confidence": 16, "Total Works": 13,
    }
    for col_idx, h in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(h, 7)

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ── Legend sheet ──────────────────────────────────────────────────────────
    lg = wb.create_sheet("Legend")
    lg["A1"] = "Color Legend — Match Confidence"
    lg["A1"].font = Font(bold=True, size=12, name="Arial")

    legend_rows = [
        (None,    WHITE,      "No highlight",  "HIGH confidence — likely correct"),
        (None,    LIGHT_BLUE, "Light Blue",    "MEDIUM confidence — verify the matched name"),
        (None,    YELLOW,     "Yellow",        "LOW confidence — verify manually"),
        (None,    ORANGE,     "Orange",        "NOT FOUND in OpenAlex"),
    ]
    for r_offset, (_, fill_hex, label, desc) in enumerate(legend_rows, 3):
        cell_a = lg.cell(row=r_offset, column=1, value=label)
        cell_a.fill = PatternFill("solid", start_color=fill_hex)
        cell_a.font = Font(name="Arial", size=10)
        lg.cell(row=r_offset, column=2, value=desc).font = Font(name="Arial", size=10)

    lg["A9"]  = "Tip: For any MEDIUM or LOW match, check the 'OpenAlex Matched Name' column."
    lg["A10"] = "      Open https://openalex.org/authors/{OpenAlex ID} to verify the author profile."
    lg["A9"].font  = Font(italic=True, name="Arial", size=10)
    lg["A10"].font = Font(italic=True, name="Arial", size=10)
    lg.column_dimensions["A"].width = 20
    lg.column_dimensions["B"].width = 55

    wb.save(OUTPUT_FILE)

    # ── Summary ───────────────────────────────────────────────────────────────
    high     = sum(1 for r in rows if r.get("Match Confidence") == "HIGH")
    medium   = sum(1 for r in rows if r.get("Match Confidence") == "MEDIUM")
    low      = sum(1 for r in rows if r.get("Match Confidence") == "LOW")
    not_found = sum(1 for r in rows if r.get("OpenAlex ID") in ("NOT_FOUND", "API_ERROR"))

    print(f"\n{'='*62}")
    print(f"  Done!  Results → '{OUTPUT_FILE}'")
    print(f"  {'─'*40}")
    print(f"  ✓  HIGH confidence :  {high:2}")
    print(f"  ~  MEDIUM confidence: {medium:2}  ← check 'OpenAlex Matched Name'")
    print(f"  ?  LOW confidence :   {low:2}  ← verify manually")
    print(f"  ✗  Not found :        {not_found:2}")
    print(f"{'='*62}\n")


if __name__ == "__main__":
    main()
